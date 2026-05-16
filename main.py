import streamlit as st
import pandas as pd
import io
from PIL import Image
from openpyxl.styles import PatternFill

# 1. 앱 설정 및 로고 설정
try:
    # 깃허브에 logo.png 파일을 업로드했을 경우 브라우저 탭 아이콘으로 사용
    img = Image.open("logo.png")
    st.set_page_config(page_title="품목 비교 자동화 앱", page_icon=img, layout="wide")
except:
    st.set_page_config(page_title="품목 비교 자동화 앱", page_icon="📊", layout="wide")

# 상단 로고 이미지 표시 (파일이 있을 경우만)
col_logo, _ = st.columns([1, 4])
with col_logo:
    try:
        st.image("logo.png", width=200)
    except:
        pass

st.title("🚀 품목 비교 자동화 앱")
st.info("한국아사히마시나리 전용 엑셀 변동 내역 분석 도구입니다.")
st.write("---")

# 2. 시트별 헤더(제목 줄) 위치 정의
sheet_headers = {
    '本機 ﾌｨｰﾀﾞｰ': 8,              # 9번째 줄
    'ｻｯｶｰﾍｯﾄﾞリスト': 6,          # 7번째 줄
    '上型クイックセットチゥス': 5,  # 6번째 줄
    '移動台': 5                     # 6번째 줄
}

# 3. 파일 업로드 UI
col1, col2 = st.columns(2)
with col1:
    file1 = st.file_uploader("첫 번째(기준: A) 파일 업로드", type=['xlsx'])
with col2:
    file2 = st.file_uploader("두 번째(비교: B) 파일 업로드", type=['xlsx'])

if file1 and file2:
    if st.button("🔍 전 시트 비교 시작"):
        xl1 = pd.ExcelFile(file1)
        xl2 = pd.ExcelFile(file2)
        
        summary_results = []
        valid_sheet_written = False
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # 시트 이름 매칭을 위한 전처리 (공백 제거)
            xl2_sheets_stripped = {str(s).strip(): s for s in xl2.sheet_names}
            
            for sheet in xl1.sheet_names:
                sheet_stripped = str(sheet).strip()
                
                if sheet_stripped in xl2_sheets_stripped:
                    sheet2_actual_name = xl2_sheets_stripped[sheet_stripped]
                    
                    header_pos = sheet_headers.get(sheet, 0)
                    df1 = xl1.parse(sheet, header=header_pos)
                    df2 = xl2.parse(sheet2_actual_name, header=header_pos)
                    
                    # 제목 정규화 (문자열 변환 및 공백 제거)
                    df1.columns = [str(c).strip() for c in df1.columns]
                    df2.columns = [str(c).strip() for c in df2.columns]
                    
                    if '品番' in df1.columns and '個数' in df1.columns:
                        valid_sheet_written = True
                        
                        # [특정 패턴 제외] '-000-'가 포함된 품번은 분석에서 제외
                        df1 = df1[~df1['品番'].astype(str).str.contains('-000-', na=False)].copy()
                        df2 = df2[~df2['品番'].astype(str).str.contains('-000-', na=False)].copy()
                        
                        # 1. 신규 추가 (B에만 있는 품번)
                        new = df2[~df2['品番'].isin(df1['品番'])].copy()
                        new['변경유형'] = '신규 추가'
                        new['個数_신규'] = new['個数']
                        
                        # 2. 개수 변경 (A, B 모두 존재하나 수량이 다른 품번)
                        merged = pd.merge(df2, df1[['品番', '個数']], on='品番', suffixes=('_신규', '_기존'))
                        changed = merged[merged['個数_신규'] != merged['個数_기존']].copy()
                        changed['변경유형'] = '개수 변경'
                        changed['個数'] = changed['個数_기존']
                        
                        # 3. 삭제 (A에는 있었으나 B에서 사라진 품번)
                        deleted = df1[~df1['品番'].isin(df2['品番'])].copy()
                        deleted['변경유형'] = '삭제'
                        deleted['個数_신규'] = '-'
                        
                        # 모든 변동 사항 합치기
                        final = pd.concat([new, changed, deleted], ignore_index=True)
                        
                        new_count = len(new)
                        change_count = len(changed)
                        delete_count = len(deleted)
                        
                        summary_results.append({
                            '시트명': sheet,
                            '신규추가': new_count,
                            '개수변경': change_count,
                            '삭제': delete_count,
                            '총변동': len(final)
                        })
                        
                        # 변동 사항이 전혀 없을 경우에도 기본 양식 시트는 생성
                        if len(final) == 0:
                            pd.DataFrame(columns=['品番', '個数', '個数_신규', '변경유형']).to_excel(writer, sheet_name=sheet, index=False)
                            continue
                        
                        # 불필요한 열 삭제 (빈 열, 기존개수, 임시열 등)
                        invalid_cols = [
                            c for c in final.columns 
                            if 'Unnamed:' in str(c) 
                            or str(c).strip() == '' 
                            or str(c) in ['기존개수', '個数_기존']
                        ]
                        final = final.drop(columns=invalid_cols)
                        
                        # 열 순서 조정: 個数 바로 옆에 個数_신규 배치
                        cols = list(final.columns)
                        if '個수' in cols and '個数_신규' in cols:
                            cols.remove('個数_신규')
                            idx = cols.index('個数')
                            cols.insert(idx + 1, '個수_신규')
                            final = final[cols]
                        
                        # 엑셀 시트에 쓰기
                        final.to_excel(writer, sheet_name=sheet, index=False)
                        
                        # 스타일 서식 적용
                        worksheet = writer.sheets[sheet]
                        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        
                        type_col_idx = None
                        new_col_idx = None
                        
                        for cell in worksheet[1]:
                            if cell.value == '변경유형': type_col_idx = cell.column
                            if cell.value == '個数_신규': new_col_idx = cell.column
                        
                        for row in range(2, worksheet.max_row + 1):
                            change_type = worksheet.cell(row=row, column=type_col_idx).value if type_col_idx else ""
                            
                            if change_type == '삭제':
                                # 삭제 행은 전체 빨간색
                                for col in range(1, worksheet.max_column + 1):
                                    worksheet.cell(row=row, column=col).fill = red_fill
                            elif change_type in ['신규 추가', '개수 변경']:
                                # 신규/변경 항목은 수량 칸만 노란색
                                if new_col_idx:
                                    worksheet.cell(row=row, column=new_col_idx).fill = yellow_fill
                        
                        # 열 너비 자동 조절
                        for col in worksheet.columns:
                            max_len = max(len(str(cell.value or '')) for cell in col)
                            col_letter = col[0].column_letter
                            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
            # 매칭 시트가 아예 없을 경우 에러 방지용 시트
            if not valid_sheet_written:
                pd.DataFrame({'알림': ['일치하는 시트가 없습니다.']}).to_excel(writer, sheet_name='안내', index=False)
        
        # 4. 웹 화면 요약 리포트 출력
        st.success("✅ 분석이 완료되었습니다!")
        st.subheader("📋 시트별 변동 내역 요약")
        
        if not summary_results:
            st.warning("파일간 일치하는 시트 이름이 없습니다.")
        else:
            for res in summary_results:
                if res['총변동'] > 0:
                    st.info(f"📄 **{res['시트명']}**: 총 {res['총변동']}건 (추가 {res['신규추가']}, 변경 {res['개수변경']}, 삭제 {res['삭제']})")
                else:
                    st.write(f"⚪ **{res['시트명']}**: 변동 사항 없음")
        
        st.write("---")
        st.download_button(label="💾 결과 엑셀 파일 다운로드", data=output.getvalue(), file_name="품목비교결과.xlsx")
